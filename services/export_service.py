import csv
import re
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from database.connection import SessionLocal
from database.models import Student, Department, Venue, TimeSlot
from config import EXPORTS_DIR
from core.exceptions import ExportError

class ExportService:
    """Enterprise Data Export Service supporting multi-sheet formatted Excel, CSV, and printable PDF reports."""

    @staticmethod
    def sanitize_cell(value: Any) -> Any:
        """Protects against Excel formula injection attacks by prepending single quote to formula symbols."""
        if isinstance(value, str):
            val = value.strip()
            if val.startswith(('=', '+', '-', '@', '\t', '\r')):
                return f"'{val}"
            return val
        return value

    @classmethod
    def _clean_group_name(cls, g_name: Optional[str]) -> str:
        if not g_name:
            return "Unassigned"
        if g_name.startswith("Group "):
            return g_name[6:]
        return g_name

    @classmethod
    def _get_slot_timings(cls, session: Session, group_name: Optional[str] = None) -> Tuple[str, str, str]:
        """Queries and sorts TimeSlots from the database to return Slot 1, Slot 2, and Slot 3 timings."""
        query = session.query(TimeSlot)
        if group_name:
            query = query.filter(TimeSlot.group_name == group_name)
        time_slots = query.all()
        
        def format_time_str(t_str):
            try:
                t_str = re.sub(r'\s+', ' ', t_str.strip().upper())
                # match hour:minute AM/PM
                match = re.match(r'(\d+):(\d+)\s*(AM|PM)', t_str)
                if match:
                    hour = int(match.group(1))
                    minute = int(match.group(2))
                    period = match.group(3)
                    return f"{hour}:{minute:02d} {period}"
                # Handle cases without minutes: "9 AM"
                match_no_min = re.match(r'(\d+)\s*(AM|PM)', t_str)
                if match_no_min:
                    hour = int(match_no_min.group(1))
                    period = match_no_min.group(2)
                    return f"{hour}:00 {period}"
            except:
                pass
            return t_str

        def parse_time_to_minutes(t_str):
            try:
                t_str = re.sub(r'\s+', ' ', t_str.strip().upper())
                match = re.match(r'(\d+):(\d+)\s*(AM|PM)', t_str)
                if match:
                    hour = int(match.group(1))
                    minute = int(match.group(2))
                    period = match.group(3)
                    if period == "PM" and hour < 12:
                        hour += 12
                    elif period == "AM" and hour == 12:
                        hour = 0
                    return hour * 60 + minute
                match_no_min = re.match(r'(\d+)\s*(AM|PM)', t_str)
                if match_no_min:
                    hour = int(match_no_min.group(1))
                    period = match_no_min.group(2)
                    if period == "PM" and hour < 12:
                        hour += 12
                    elif period == "AM" and hour == 12:
                        hour = 0
                    return hour * 60
            except:
                pass
            return 0

        sorted_slots = sorted(time_slots, key=lambda ts: (ts.day_number, parse_time_to_minutes(ts.start_time), ts.slot_name or "", ts.id))
        
        slot1_val = ""
        slot2_val = ""
        slot3_val = ""
        if len(sorted_slots) >= 1:
            slot1_val = f"{format_time_str(sorted_slots[0].start_time)} - {format_time_str(sorted_slots[0].end_time)}"
        if len(sorted_slots) >= 2:
            slot2_val = f"{format_time_str(sorted_slots[1].start_time)} - {format_time_str(sorted_slots[1].end_time)}"
        if len(sorted_slots) >= 3:
            slot3_val = f"{format_time_str(sorted_slots[2].start_time)} - {format_time_str(sorted_slots[2].end_time)}"
            
        return slot1_val, slot2_val, slot3_val

    @classmethod
    def export_group_wise_excel(cls, destination_path: Path) -> Path:
        """Generates Group-wise Allocation Excel file in the standard layout."""
        session: Session = SessionLocal()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Group-wise Allocation"

            students = session.query(Student).filter(Student.is_deleted == False).order_by(Student.usn.asc()).all()
            
            slots_by_group = {
                "Group A": cls._get_slot_timings(session, "Group A"),
                "Group B": cls._get_slot_timings(session, "Group B"),
                "Unassigned": cls._get_slot_timings(session, None)
            }

            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

            columns = ["Sl No.", "Student ID", "USN", "Name", "Branch", "Group", "Slot 1", "Slot 2", "Slot 3", "Venue"]
            ws.append(columns)

            for col_num in range(1, len(columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for idx, s in enumerate(students, 1):
                branch_val = ""
                if s.department:
                    branch_val = s.department.code or s.department.name
                
                venue_val = "Unassigned"
                if s.group_venue:
                    venue_val = s.group_venue.name
                elif s.venue:
                    venue_val = s.venue.name

                g_name = s.group_name or "Unassigned"
                if g_name not in slots_by_group:
                    slots_by_group[g_name] = cls._get_slot_timings(session, g_name)

                slot1_val, slot2_val, slot3_val = slots_by_group[g_name]

                ws.append([
                    idx,
                    cls.sanitize_cell(s.student_id or ""),
                    cls.sanitize_cell(s.usn),
                    cls.sanitize_cell(s.full_name),
                    cls.sanitize_cell(branch_val),
                    cls.sanitize_cell(cls._clean_group_name(s.group_name)),
                    cls.sanitize_cell(slot1_val),
                    cls.sanitize_cell(slot2_val),
                    cls.sanitize_cell(slot3_val),
                    cls.sanitize_cell(venue_val)
                ])

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            wb.save(destination_path)
            return destination_path
        except Exception as e:
            raise ExportError(f"Group-wise Excel export failed: {str(e)}")
        finally:
            session.close()

    @classmethod
    def export_branch_wise_excel(cls, destination_path: Path) -> Path:
        """Generates Branch-wise Allocation Excel file in the standard layout."""
        session: Session = SessionLocal()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Branch-wise Allocation"

            students = session.query(Student).filter(Student.is_deleted == False).order_by(Student.usn.asc()).all()
            slot1, slot2, slot3 = cls._get_slot_timings(session)

            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

            columns = ["Sl No.", "Student ID", "USN", "Name", "Branch", "Group", "Slot 1", "Slot 2", "Slot 3", "Venue"]
            ws.append(columns)

            for col_num in range(1, len(columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for idx, s in enumerate(students, 1):
                branch_val = ""
                if s.department:
                    branch_val = s.department.code or s.department.name

                venue_val = "Unassigned"
                if s.branch_venue:
                    venue_val = s.branch_venue.name
                elif s.venue:
                    venue_val = s.venue.name

                ws.append([
                    idx,
                    cls.sanitize_cell(s.student_id or ""),
                    cls.sanitize_cell(s.usn),
                    cls.sanitize_cell(s.full_name),
                    cls.sanitize_cell(branch_val),
                    cls.sanitize_cell(cls._clean_group_name(s.group_name)),
                    cls.sanitize_cell(slot1),
                    cls.sanitize_cell(slot2),
                    cls.sanitize_cell(slot3),
                    cls.sanitize_cell(venue_val)
                ])

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            wb.save(destination_path)
            return destination_path
        except Exception as e:
            raise ExportError(f"Branch-wise Excel export failed: {str(e)}")
        finally:
            session.close()

    @classmethod
    def export_excel_master(cls, destination_path: Path) -> Path:
        """Generates a styled Master Excel workbook in the standard layout."""
        session: Session = SessionLocal()
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            students = session.query(Student).filter(Student.is_deleted == False).order_by(Student.usn.asc()).all()
            
            slots_by_group = {
                "Group A": cls._get_slot_timings(session, "Group A"),
                "Group B": cls._get_slot_timings(session, "Group B"),
                "Unassigned": cls._get_slot_timings(session, None)
            }

            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

            columns = ["Sl No.", "Student ID", "USN", "Name", "Branch", "Group", "Slot 1", "Slot 2", "Slot 3", "Venue"]

            ws_master = wb.create_sheet(title="Master Allocation")
            ws_master.append(columns)

            for col_num in range(1, len(columns) + 1):
                cell = ws_master.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for idx, s in enumerate(students, 1):
                branch_val = ""
                if s.department:
                    branch_val = s.department.code or s.department.name

                venue_val = "Unassigned"
                if s.group_venue:
                    venue_val = s.group_venue.name
                elif s.branch_venue:
                    venue_val = s.branch_venue.name
                elif s.venue:
                    venue_val = s.venue.name

                g_name = s.group_name or "Unassigned"
                if g_name not in slots_by_group:
                    slots_by_group[g_name] = cls._get_slot_timings(session, g_name)

                slot1_val, slot2_val, slot3_val = slots_by_group[g_name]

                row = [
                    idx,
                    cls.sanitize_cell(s.student_id or ""),
                    cls.sanitize_cell(s.usn),
                    cls.sanitize_cell(s.full_name),
                    cls.sanitize_cell(branch_val),
                    cls.sanitize_cell(cls._clean_group_name(s.group_name)),
                    cls.sanitize_cell(slot1_val),
                    cls.sanitize_cell(slot2_val),
                    cls.sanitize_cell(slot3_val),
                    cls.sanitize_cell(venue_val)
                ]
                ws_master.append(row)

            for sheet in wb.worksheets:
                for col in sheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

            wb.save(destination_path)
            return destination_path
        except Exception as e:
            raise ExportError(f"Excel export failed: {str(e)}")
        finally:
            session.close()

    @classmethod
    def export_csv(cls, destination_path: Path) -> Path:
        """Generates flat CSV export matching the standard columns."""
        session: Session = SessionLocal()
        try:
            students = session.query(Student).filter(Student.is_deleted == False).order_by(Student.usn.asc()).all()
            
            slots_by_group = {
                "Group A": cls._get_slot_timings(session, "Group A"),
                "Group B": cls._get_slot_timings(session, "Group B"),
                "Unassigned": cls._get_slot_timings(session, None)
            }

            headers = ["Sl No.", "Student ID", "USN", "Name", "Branch", "Group", "Slot 1", "Slot 2", "Slot 3", "Venue"]

            with open(destination_path, mode="w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)

                for idx, s in enumerate(students, 1):
                    branch_val = ""
                    if s.department:
                        branch_val = s.department.code or s.department.name

                    venue_val = "Unassigned"
                    if s.group_venue:
                        venue_val = s.group_venue.name
                    elif s.branch_venue:
                        venue_val = s.branch_venue.name
                    elif s.venue:
                        venue_val = s.venue.name

                    g_name = s.group_name or "Unassigned"
                    if g_name not in slots_by_group:
                        slots_by_group[g_name] = cls._get_slot_timings(session, g_name)

                    slot1_val, slot2_val, slot3_val = slots_by_group[g_name]

                    writer.writerow([
                        idx,
                        cls.sanitize_cell(s.student_id or ""),
                        cls.sanitize_cell(s.usn),
                        cls.sanitize_cell(s.full_name),
                        cls.sanitize_cell(branch_val),
                        cls.sanitize_cell(cls._clean_group_name(s.group_name)),
                        cls.sanitize_cell(slot1_val),
                        cls.sanitize_cell(slot2_val),
                        cls.sanitize_cell(slot3_val),
                        cls.sanitize_cell(venue_val)
                    ])

            return destination_path
        except Exception as e:
            raise ExportError(f"CSV export failed: {str(e)}")
        finally:
            session.close()

    @classmethod
    def export_pdf_attendance_sheet(cls, destination_path: Path, venue_id: Optional[int] = None) -> Path:
        """Generates a professional PDF attendance sheet for venue/faculty use."""
        session: Session = SessionLocal()
        try:
            query = session.query(Student).filter(Student.is_deleted == False)
            if venue_id:
                query = query.filter(Student.venue_id == venue_id)
            
            students = query.order_by(Student.group_name.asc(), Student.department_id.asc(), Student.full_name.asc()).all()

            doc = SimpleDocTemplate(
                str(destination_path),
                pagesize=A4,
                rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'DocTitle',
                parent=styles['Heading1'],
                fontSize=18,
                leading=22,
                textColor=colors.HexColor('#1E293B'),
                alignment=1
            )
            subtitle_style = ParagraphStyle(
                'DocSubTitle',
                parent=styles['Normal'],
                fontSize=10,
                leading=12,
                textColor=colors.HexColor('#64748B'),
                alignment=1
            )

            elements = []
            elements.append(Paragraph("COLLEGE INDUCTION PROGRAM - ATTENDANCE SHEET", title_style))
            elements.append(Spacer(1, 4))
            elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Students: {len(students)}", subtitle_style))
            elements.append(Spacer(1, 15))

            table_data = [["S.No", "USN", "Student Name", "Dept", "Group", "Venue / Slot", "Signature"]]

            for idx, s in enumerate(students, 1):
                venue_slot_str = f"{s.venue.name if s.venue else 'N/A'}\n({s.time_slot.slot_name if s.time_slot else 'N/A'})"
                table_data.append([
                    str(idx),
                    s.usn,
                    s.full_name,
                    s.department.name[:15] if s.department else "",
                    s.group_name or "",
                    venue_slot_str,
                    "[  ] Present"
                ])

            t = Table(table_data, colWidths=[30, 85, 140, 75, 50, 95, 75])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elements.append(t)
            doc.build(elements)
            return destination_path

        except Exception as e:
            raise ExportError(f"PDF export failed: {str(e)}")
        finally:
            session.close()
